import xml.etree.ElementTree
import pandas as pd
import math
import sys

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.worksheet.dimensions import ColumnDimension

class XmlConfigParserException(Exception):
    pass

class XmlConfigParser():
    def __init__(self, xmlfile):
        self.cfg = {}
        try:
            self.root = xml.etree.ElementTree.parse(xmlfile).getroot() 
        except xml.etree.ElementTree.ParseError:
            self.root = None
        
        if self.root is None:
            raise XmlConfigParserException("Kan configuratie niet lezen")

        for elem in self.root.iter():
            self.cfg[elem.tag] = elem.text

    def get(self, key, default=""):
        if key in self.cfg:
            return self.cfg[key]
        else:
            return default
 
class Calculator():
    CONST_JAAR                      = 1
    CONST_BESPARING                 = 7
    CONST_JAARLIJKSE_SUBSIDIE       = 8
    CONST_EENMALIGE_KOSTEN          = 11
    CONST_VASTE_EXPLOITATIEKOSTEN   = 12
    CONST_HERINVESTERING            = 13
    CONST_EBITDA                    = 15
    CONST_AFSCHRIJVINGSKOSTEN       = 17
    CONST_FINANCIERINGSKOSTEN       = 18
    CONST_BELASTING                 = 19
    CONST_WINSTNABELASTING          = 21
    CONST_CASHFLOW_IRR              = 22
    CONST_CASHFLOW_REV              = 23
    CONST_TVT                       = 24

    def __init__(self, config):
        self.config = config
        self.result = pd.DataFrame()
        self.irr = 0.0
        self.rev = 0.0
        self.winst = 0.0
        self.tvt = 0.0


    def berekenIRR(self, cash_flows, guess=0.1, max_iterations=1000, tolerance=1e-6):
        def npv(rate):
            return sum(cf / (1 + rate) ** i for i, cf in enumerate(cash_flows))

        def derivative(rate):
            return sum(-i * cf / (1 + rate) ** (i + 1) for i, cf in enumerate(cash_flows))

        rate = guess
        for _ in range(max_iterations):
            value = npv(rate)
            deriv = derivative(rate)
            if deriv == 0:
                raise ZeroDivisionError("Derivative is zero; try another guess.")
            new_rate = rate - value / deriv
            if abs(new_rate - rate) < tolerance:
                return new_rate
            rate = new_rate
        raise RuntimeError("IRR did not converge")

    def doCalculations(self):
        self.result = pd.DataFrame(columns=range(int(self.config.get("Termijn", "12")) + 2),index=range(25))

        aflossing=((float(self.config.get("Investering", "160000")) - float(self.config.get("EenmaligeSubsidie", "10000"))) * (1 - float(self.config.get("EigenVermogen", "0.2")))) / float(self.config.get("Termijn", "12"))
        totaleinvestering=float(self.config.get("Investering", "160000")) - float(self.config.get("EenmaligeSubsidie", "10000"))
        afschrijving=(float(self.config.get("Investering", "160000")) - float(self.config.get("EenmaligeSubsidie", "10000")) - float(self.config.get("Restwaarde", "0"))) / float(self.config.get("Termijn", "12"))

        irr = []
        rev = []

        for jaar in range(int(self.config.get("Termijn", "12")) + 1):
            self.result.iat[self.CONST_JAAR,jaar+1] = jaar
            if jaar == 0:
                #Vul kosten voor jaar 0
                self.result.iat[self.CONST_EENMALIGE_KOSTEN,jaar+1] = -1 * float(self.config.get("EenmaligeKosten", "30000"))
                self.result.iat[self.CONST_VASTE_EXPLOITATIEKOSTEN,jaar+1] = 0
                self.result.iat[self.CONST_HERINVESTERING,jaar+1] = 0
                self.result.iat[self.CONST_EBITDA,jaar+1] = -1 * float(self.config.get("EenmaligeKosten", "30000"))
                self.result.iat[self.CONST_AFSCHRIJVINGSKOSTEN,jaar+1] = 0
                self.result.iat[self.CONST_FINANCIERINGSKOSTEN,jaar+1] = 0
                self.result.iat[self.CONST_BELASTING,jaar+1] = 0
                self.result.iat[self.CONST_WINSTNABELASTING,jaar+1] = -1 * float(self.config.get("EenmaligeKosten", "30000"))
                self.winst += self.result.iat[self.CONST_WINSTNABELASTING,jaar+1]
                self.result.iat[self.CONST_CASHFLOW_IRR,jaar+1] =  -1 * (float(self.config.get("EenmaligeKosten", "30000")) + totaleinvestering)
                self.result.iat[self.CONST_CASHFLOW_REV,jaar+1] =  -1 * (((totaleinvestering) * float(self.config.get("EigenVermogen", "0.2"))) - self.result.iat[self.CONST_WINSTNABELASTING,jaar+1])
                irr.append(self.result.iat[self.CONST_CASHFLOW_IRR,jaar+1])
                rev.append(self.result.iat[self.CONST_CASHFLOW_REV,jaar+1])
            else:
                self.result.iat[self.CONST_BESPARING,jaar+1] = float(self.config.get("Besparing", "20000")) * ((1 + float(self.config.get("Inflatie", "0.02"))) ** (jaar - 1))
                self.result.iat[self.CONST_JAARLIJKSE_SUBSIDIE,jaar+1] = float(self.config.get("JaarlijkseSubsidie", "800")) * ((1 + float(self.config.get("Inflatie", "0.02"))) ** (jaar - 1))
                self.result.iat[self.CONST_EENMALIGE_KOSTEN,jaar+1] = 0
                self.result.iat[self.CONST_VASTE_EXPLOITATIEKOSTEN,jaar+1] = -1 * (float(self.config.get("VasteExploitatieKosten", "2000")) * ((1 + float(self.config.get("Inflatie", "0.02"))) ** (jaar - 1)))
                if  jaar != int(self.config.get("HerinvesteringJaar", "6")):
                    self.result.iat[self.CONST_HERINVESTERING,jaar+1] = 0
                else:
                    self.result.iat[self.CONST_HERINVESTERING,jaar+1] = -1 * float(self.config.get("Herinvestering", "4000"))
                self.result.iat[self.CONST_EBITDA,jaar+1] = self.result.iat[self.CONST_BESPARING,jaar+1] + self.result.iat[self.CONST_JAARLIJKSE_SUBSIDIE,jaar+1] + self.result.iat[self.CONST_VASTE_EXPLOITATIEKOSTEN,jaar+1] + self.result.iat[self.CONST_HERINVESTERING,jaar+1]
                self.result.iat[self.CONST_AFSCHRIJVINGSKOSTEN,jaar+1] = -1 * afschrijving
                self.result.iat[self.CONST_FINANCIERINGSKOSTEN,jaar+1] = -1 * ((totaleinvestering * (1 - float(self.config.get("EigenVermogen", "0.2")))) - (aflossing * (jaar - 1))) * float(self.config.get("RenteVV", "0.04"))
                winstvoorbelasting=self.result.iat[self.CONST_EBITDA,jaar+1] + self.result.iat[self.CONST_AFSCHRIJVINGSKOSTEN,jaar+1] + self.result.iat[self.CONST_FINANCIERINGSKOSTEN,jaar+1]
                if  winstvoorbelasting > 0:
                    self.result.iat[self.CONST_BELASTING,jaar+1] = -1 * (winstvoorbelasting * float(self.config.get("Belasting", "0.165")))
                else:
                    self.result.iat[self.CONST_BELASTING,jaar+1] = 0
                self.result.iat[self.CONST_WINSTNABELASTING,jaar+1] = winstvoorbelasting + self.result.iat[self.CONST_BELASTING,jaar+1]
                self.winst += self.result.iat[self.CONST_WINSTNABELASTING,jaar+1]
                self.result.iat[self.CONST_CASHFLOW_IRR,jaar+1] = self.result.iat[self.CONST_EBITDA,jaar+1] + self.result.iat[self.CONST_BELASTING,jaar+1]
                self.result.iat[self.CONST_CASHFLOW_REV,jaar+1] = self.result.iat[self.CONST_EBITDA,jaar+1] + self.result.iat[self.CONST_FINANCIERINGSKOSTEN,jaar+1] + self.result.iat[self.CONST_BELASTING,jaar+1] - aflossing
                tvttotaal = 0
                for tvtjaar in range(jaar + 1):
                    tvttotaal += self.result.iat[self.CONST_CASHFLOW_IRR,tvtjaar + 1]
                if tvttotaal > 0:
                    self.result.iat[self.CONST_TVT,jaar+1] = jaar - (self.result.iat[self.CONST_CASHFLOW_REV,jaar+1] / self.result.iat[self.CONST_CASHFLOW_IRR,jaar+1])
                    if self.tvt == 0:
                        self.tvt = self.result.iat[self.CONST_TVT,jaar+1]
                irr.append(self.result.iat[self.CONST_CASHFLOW_IRR,jaar+1])
                rev.append(self.result.iat[self.CONST_CASHFLOW_REV,jaar+1])
        self.irr = self.berekenIRR(irr)
        self.rev = self.berekenIRR(rev)

    def writeOutput(self):
        workbook = load_workbook('template.xlsx')
        worksheet = workbook['Opdracht 2']
        output = workbook['Output']
        
        start_row = 1
        start_col = 1
        for r_idx, row in enumerate(dataframe_to_rows(self.result, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                if not math.isnan(value):
                    output.cell(row=r_idx, column=c_idx, value=value)

        # Input gegevens
        worksheet.cell(row=4, column=2, value=float(self.config.get("Inflatie", "0.02")))
        worksheet.cell(row=5, column=2, value=self.config.get("Afschrijving", "Linear"))
        worksheet.cell(row=6, column=2, value=float(self.config.get("EigenVermogen", "0.2")))
        worksheet.cell(row=7, column=2, value=float(self.config.get("RenteVV", "0.04")))
        worksheet.cell(row=8, column=2, value=float(self.config.get("Belasting", "0.165")))

        worksheet.cell(row=12, column=2, value=float(self.config.get("Termijn", "12")))
        worksheet.cell(row=13, column=2, value=float(self.config.get("HerinvesteringJaar", "6")))

        worksheet.cell(row=15, column=2, value=float(self.config.get("Investering", "160000")))
        worksheet.cell(row=16, column=2, value=float(self.config.get("EenmaligeSubsidie", "10000")))
        worksheet.cell(row=17, column=2, value=float(self.config.get("Restwaarde", "0")))

        worksheet.cell(row=20, column=2, value=float(self.config.get("Besparing", "20000")))
        worksheet.cell(row=21, column=2, value=float(self.config.get("JaarlijkseSubsidie", "800")))

        worksheet.cell(row=24, column=2, value=float(self.config.get("EenmaligeKosten", "30000")))
        worksheet.cell(row=25, column=2, value=float(self.config.get("VasteExploitatieKosten", "2000")))
        worksheet.cell(row=26, column=2, value=float(self.config.get("Herinvestering", "4000")))

        worksheet.cell(row=29, column=2, value=float(self.config.get("Investering", "160000")) - float(self.config.get("EenmaligeSubsidie", "10000")))
        worksheet.cell(row=30, column=2, value=(float(self.config.get("Investering", "160000")) - float(self.config.get("EenmaligeSubsidie", "10000")) - float(self.config.get("Restwaarde", "0"))) / float(self.config.get("Termijn", "12")))
        worksheet.cell(row=31, column=2, value=((float(self.config.get("Investering", "160000")) - float(self.config.get("EenmaligeSubsidie", "10000"))) * (1 - float(self.config.get("EigenVermogen", "0.2")))) / float(self.config.get("Termijn", "12")))

        worksheet.cell(row=36, column=2, value=self.irr)
        worksheet.cell(row=37, column=2, value=self.rev)
        worksheet.cell(row=38, column=2, value=self.winst)
        worksheet.cell(row=39, column=2, value=self.tvt)

        # Autofit columns in de output worksheet
        for column in output.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    curr_length = len(str(cell.value))
                    if curr_length > max_length:
                        max_length = curr_length
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            output.column_dimensions[column_letter].width = adjusted_width

        workbook.save(self.config.get("Outputfile", "Opdracht2.xlsx"))

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Gebruik: python opdracht2.py config.xml")
        sys.exit(1)

    filename = sys.argv[1]

    config = XmlConfigParser(filename)
    calculator = Calculator(config)
    calculator.doCalculations()
    calculator.writeOutput()
