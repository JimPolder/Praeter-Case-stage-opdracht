import xml.etree.ElementTree
import pandas as pd
import math
import sys

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, Series

class XmlConfigParserException(Exception):
    pass

class XmlConfigParser():
    def __init__(self, xmlfile):
        self.cfg = {}
        try:
            self.root = xml.etree.ElementTree.parse(xmlfile).getroot() 
        except xml.etree.ElementTree.ParseError:
            self.root = None
        
        if self.root is None:
            raise XmlConfigParserException("Kan configuratie niet lezen")

        for elem in self.root.iter():
            self.cfg[elem.tag] = elem.text

    def get(self, key, default=""):
        if key in self.cfg:
            return self.cfg[key]
        else:
            return default
 
class Calculator():
    def __init__(self, config):
        self.config = config
        self.cbsdata = pd.DataFrame()
        self.result = pd.DataFrame(columns=range(35),index=range(35))

    def readCBS(self):
        filename = self.config.get("CBSDatafile", "cbsdata.xlsx")
        self.cbsdata = pd.read_excel(filename)

    def doCalculations(self):
        # Zoek categorie
        query = 'Categorie == "' + self.config.get('Categorie','Detailhandel zonder koeling') + '" and From <= ' + self.config.get('Bouwjaar','2025') + ' and To >= ' + self.config.get('Bouwjaar','2025')
        categorie = self.cbsdata.query(query)
        # Zoek gemiddelde gas en elektriciteits gebruik op basis van de gegeven oppervlakte
        opp = [ 250, 500, 1000, 2500, 5000 ]
        oppindex = 0
        for _opp in opp:
            if _opp >= int(self.config.get("Oppervlakte", "5000")):
                break
            oppindex+=1
        gemgasgebruik = float(categorie.iat[0, oppindex + 4])
        gemelektriciteitsgebruik = float(categorie.iat[0, oppindex + 9])
        # Bereken het huidige gas en elektriciteits gebruik per m2 
        huidiggasgebruik = float(self.config.get("Gas", "130000")) / float(self.config.get("Oppervlakte", "5000"))
        huidigelektriciteitsgebruik = float(self.config.get("Elektriciteit", "27400")) / float(self.config.get("Oppervlakte", "5000"))

        # Huidig verbruik per m2
        self.result.iat[5,9] = huidiggasgebruik
        self.result.iat[6,9] = huidigelektriciteitsgebruik
        self.result.iat[7,9] = huidiggasgebruik * float(self.config.get("EnergetischeWaardeGasElektra", "10")) + huidigelektriciteitsgebruik
        
        # Huidig verbruik per m3
        self.result.iat[5,10] = huidiggasgebruik / float(self.config.get("HoogtePlafond", "2.7"))
        self.result.iat[6,10] = huidigelektriciteitsgebruik / float(self.config.get("HoogtePlafond", "2.7"))
        self.result.iat[7,10] = (huidiggasgebruik * float(self.config.get("EnergetischeWaardeGasElektra", "10")) + huidigelektriciteitsgebruik) / float(self.config.get("HoogtePlafond", "2.7"))
        
        # Huidig verbruik totaal
        self.result.iat[5,11] = huidiggasgebruik * float(self.config.get("Oppervlakte", "5000"))
        self.result.iat[6,11] = huidigelektriciteitsgebruik * float(self.config.get("Oppervlakte", "5000"))
        self.result.iat[7,11] = (huidiggasgebruik * float(self.config.get("EnergetischeWaardeGasElektra", "10")) + huidigelektriciteitsgebruik) * float(self.config.get("Oppervlakte", "5000"))
        
        # Gemiddeld verbruik per m2
        self.result.iat[10,9] = gemgasgebruik
        self.result.iat[11,9] = gemelektriciteitsgebruik
        self.result.iat[12,9] = gemgasgebruik * float(self.config.get("EnergetischeWaardeGasElektra", "10")) + gemelektriciteitsgebruik
        
        # Gemiddeld verbruik per m3
        self.result.iat[10,10] = gemgasgebruik / float(self.config.get("GemHoogtePlafond", "2.7"))
        self.result.iat[11,10] = gemelektriciteitsgebruik / float(self.config.get("GemHoogtePlafond", "2.7"))
        self.result.iat[12,10] = (gemgasgebruik * float(self.config.get("EnergetischeWaardeGasElektra", "10")) + gemelektriciteitsgebruik) / float(self.config.get("GemHoogtePlafond", "2.7"))
        
        # Gemiddeld verbruik totaal
        self.result.iat[10,11] = gemgasgebruik * float(self.config.get("Oppervlakte", "5000"))
        self.result.iat[11,11] = gemelektriciteitsgebruik * float(self.config.get("Oppervlakte", "5000"))
        self.result.iat[12,11] = (gemgasgebruik * float(self.config.get("EnergetischeWaardeGasElektra", "10")) + gemelektriciteitsgebruik) * float(self.config.get("Oppervlakte", "5000"))

        # Totaal (factor)
        self.result.iat[15,3] = float(self.config.get("Elektriciteit", "27400")) * float(self.config.get("EnergetischeWaardeGasElektra", "10"))

        # Data voor grafiek
        self.result.iat[1,30] = self.result.iat[7,9]
        self.result.iat[1,31] = self.result.iat[12,9]
        self.result.iat[2,30] = self.result.iat[7,10]
        self.result.iat[2,31] = self.result.iat[12,10]
        
        
    def writeOutput(self):
        workbook = load_workbook('template.xlsx')
        worksheet = workbook['Opdracht 1']
        
        start_row = 1
        start_col = 1
        for r_idx, row in enumerate(dataframe_to_rows(self.result, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                if not math.isnan(value):
                    worksheet.cell(row=r_idx, column=c_idx, value=value)

        # Grafiek toevoegen
        worksheet.cell(row=1,column=31, value="Huidig verbruik")
        worksheet.cell(row=1,column=32, value="Gemiddeld verbruik")
        worksheet.cell(row=2,column=30, value="per m2")
        worksheet.cell(row=3,column=30, value="per m3")

        data = Reference(worksheet, min_col=31, max_col=32, min_row=1, max_row=3)
        categories = Reference(worksheet, min_col=30, max_col=30, min_row=2, max_row=3)

        chart = BarChart()
        chart.title = "Huidig verbruik t.o.v. gemiddeld verbruik (kWh)"
        chart.y_axis.delete = False
        chart.x_axis.delete = False

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        worksheet.add_chart(chart, "P4")
        
        # Input gegevens
        worksheet.cell(row=6, column=4, value=self.config.get("Naam", "Bedrijf X"))
        worksheet.cell(row=7, column=4, value=self.config.get("Straat", "Nassau Ouwerkerkstraat 3"))
        worksheet.cell(row=8, column=4, value=self.config.get("Postcode", "2596CC"))
        worksheet.cell(row=9, column=4, value=self.config.get("Plaats", "Den Haag"))

        worksheet.cell(row=13, column=4, value=int(self.config.get("Gas", "130000")))
        worksheet.cell(row=14, column=4, value=int(self.config.get("Elektriciteit", "27400")))
        worksheet.cell(row=15, column=4, value=int(self.config.get("EnergetischeWaardeGasElektra", "10")))

        worksheet.cell(row=19, column=4, value=int(self.config.get("Verdiepingen", "3")))
        worksheet.cell(row=20, column=4, value=int(self.config.get("Bouwjaar", "2025")))
        worksheet.cell(row=21, column=4, value=self.config.get("Categorie", "Detailhandel zonder koeling"))
        worksheet.cell(row=22, column=4, value=int(self.config.get("Oppervlakte", "5000")))
        worksheet.cell(row=23, column=4, value=int(self.config.get("HoogtePlafond", "3")))

        workbook.save(self.config.get("Outputfile", "Opdracht1.xlsx"))

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Gebruik: python opdracht1.py config.xml")
        sys.exit(1)

    filename = sys.argv[1]

    config = XmlConfigParser(filename)
    calculator = Calculator(config)
    calculator.readCBS()
    calculator.doCalculations()
    calculator.writeOutput()
